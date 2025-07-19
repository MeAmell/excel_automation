import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import io
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
import os

st.title("📅 Dynamic Client Renewal Checker with Auto-Updating Worksheet")

# Upload file
uploaded_file = st.file_uploader("Upload Excel file (.xlsx)", type=["xlsx"])

def find_date_columns(df):
    """Find columns that might contain dates (birthday, renewal dates, etc.)"""
    date_columns = []
    for col in df.columns:
        col_lower = str(col).lower()
        # Check for common date column names
        if any(keyword in col_lower for keyword in ['birthday', 'dob', 'birth', 'date', 'renewal', 'premium']):
            # Try to convert sample data to datetime
            try:
                sample_data = df[col].dropna().head(10)
                if len(sample_data) > 0:
                    # For birthday columns, be more flexible with date parsing
                    if 'birthday' in col_lower or 'birth' in col_lower or 'dob' in col_lower:
                        pd.to_datetime(sample_data, errors='raise', dayfirst=True)
                    else:
                        pd.to_datetime(sample_data, errors='raise')
                    date_columns.append(col)
            except:
                # If automatic detection fails, still include if column name suggests it's a date
                if any(keyword in col_lower for keyword in ['birthday', 'dob', 'birth']):
                    date_columns.append(col)
                continue
    return date_columns

def find_email_columns(df):
    """Find columns that might contain email addresses"""
    email_columns = []
    for col in df.columns:
        col_lower = str(col).lower()
        if 'email' in col_lower or 'mail' in col_lower:
            email_columns.append(col)
    return email_columns

def detect_data_issues(df, birthday_col):
    """Detect common data issues in birthday column"""
    issues = []
    
    # Convert to datetime for analysis
    birthday_series = pd.to_datetime(df[birthday_col], errors='coerce')
    
    # Check for 1970-01-01 dates (common data issue)
    epoch_dates = birthday_series[
        (birthday_series.dt.year == 1970) & 
        (birthday_series.dt.month == 1) & 
        (birthday_series.dt.day == 1)
    ]
    
    if len(epoch_dates) > 0:
        issues.append(f"Found {len(epoch_dates)} entries with 1970-01-01 dates (likely data corruption)")
    
    # Check for future dates
    future_dates = birthday_series[birthday_series > datetime.now()]
    if len(future_dates) > 0:
        issues.append(f"Found {len(future_dates)} future birth dates")
    
    # Check for very old dates (before 1900)
    old_dates = birthday_series[birthday_series < datetime(1900, 1, 1)]
    if len(old_dates) > 0:
        issues.append(f"Found {len(old_dates)} birth dates before 1900")
    
    # Check for invalid/missing dates
    invalid_dates = df[birthday_col].isna().sum()
    if invalid_dates > 0:
        issues.append(f"Found {invalid_dates} invalid/missing birth dates")
    
    return issues

def calculate_age(birth_date):
    """Calculate age from birth date"""
    if pd.isna(birth_date):
        return None
    
    # Ensure birth_date is a datetime object
    if isinstance(birth_date, str):
        try:
            birth_date = pd.to_datetime(birth_date)
        except:
            return None
    
    today = datetime.now()
    
    # Check if birth_date is valid (not 1970-01-01 which indicates missing data)
    if birth_date.year == 1970 and birth_date.month == 1 and birth_date.day == 1:
        return None
    
    # Calculate age
    age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
    
    # Sanity check - age should be reasonable (0-120)
    if age < 0 or age > 120:
        return None
    
    return age

def get_column_letter(col_num):
    """Convert column number to Excel column letter"""
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(col_num % 26 + ord('A')) + result
        col_num //= 26
    return result

def create_task1_worksheet(wb, raw_data_sheet, birthday_col_name):
    """Create Task 1: Auto-updating worksheet for clients above 25"""
    
    # Remove existing sheet if it exists
    if "Task 1 - Clients Above 25" in wb.sheetnames:
        wb.remove(wb["Task 1 - Clients Above 25"])
    
    # Create new worksheet
    ws_new = wb.create_sheet("Task 1 - Clients Above 25")
    
    # Get raw data sheet
    ws_raw = wb[raw_data_sheet]
    
    # Find the last row and column in raw data
    max_row = ws_raw.max_row
    max_col = ws_raw.max_column
    
    # Get headers from raw data
    headers = []
    for col in range(1, max_col + 1):
        cell_value = ws_raw.cell(row=1, column=col).value
        headers.append(cell_value if cell_value else f"Column_{col}")
    
    # Add "Age" column
    headers.append("Age")
    
    # Write headers to new sheet
    for col, header in enumerate(headers, 1):
        cell = ws_new.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Find birthday column index in raw data
    birthday_col_idx = None
    for i, header in enumerate(headers[:-1]):  # Exclude Age column
        if header and str(header).lower().strip() == birthday_col_name.lower().strip():
            birthday_col_idx = i + 1
            break
    
    if not birthday_col_idx:
        print(f"Birthday column '{birthday_col_name}' not found in raw data")
        return None
    
    # Create formulas for filtering clients above 25
    for row in range(2, max_row + 1):
        # Reference to birthday cell in raw data sheet
        birthday_cell_ref = f"'{raw_data_sheet}'!{get_column_letter(birthday_col_idx)}{row}"
        
        # Age calculation formula - using DATEDIF with proper error handling
        age_formula = f'IF(ISBLANK({birthday_cell_ref}),"",IF(ISERROR(DATEDIF({birthday_cell_ref},TODAY(),"Y")),"",DATEDIF({birthday_cell_ref},TODAY(),"Y")))'
        
        # Check if age > 25 with proper error handling
        condition = f'IF(ISBLANK({birthday_cell_ref}),FALSE,IF(ISERROR(DATEDIF({birthday_cell_ref},TODAY(),"Y")),FALSE,DATEDIF({birthday_cell_ref},TODAY(),"Y")>25))'
        
        # Create conditional formulas for each column
        for col in range(1, len(headers) + 1):
            if col == len(headers):  # Age column (last column)
                formula = f'=IF({condition},{age_formula},"")'
            else:
                source_cell_ref = f"'{raw_data_sheet}'!{get_column_letter(col)}{row}"
                formula = f'=IF({condition},{source_cell_ref},"")'
            
            ws_new.cell(row=row, column=col, value=formula)
    
    # Apply table formatting
    table_ref = f"A1:{get_column_letter(len(headers))}{max_row}"
    table = Table(displayName="ClientsAbove25", ref=table_ref)
    
    # Add a default style
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    table.tableStyleInfo = style
    ws_new.add_table(table)
    
    # Auto-adjust column widths
    for column in ws_new.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_new.column_dimensions[column_letter].width = adjusted_width
    
    return ws_new

def get_column_letter(col_num):
    """Convert column number to Excel column letter"""
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(col_num % 26 + ord('A')) + result
        col_num //= 26
    return result

# Alternative approach - Create static filtered data instead of formulas
def create_task1_worksheet_static(wb, df, raw_data_sheet, birthday_col_name):
    """Create Task 1: Static worksheet for clients above 25 (alternative approach)"""
    
    # Remove existing sheet if it exists
    if "Task 1 - Clients Above 25" in wb.sheetnames:
        wb.remove(wb["Task 1 - Clients Above 25"])
    
    # Create new worksheet
    ws_new = wb.create_sheet("Task 1 - Clients Above 25")
    
    # Convert birthday column to datetime and calculate age
    df[birthday_col_name] = pd.to_datetime(df[birthday_col_name], errors='coerce', dayfirst=True)
    
    def calculate_age_safe(birth_date):
        if pd.isna(birth_date):
            return None
        today = datetime.now()
        try:
            age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
            return age if 0 <= age <= 120 else None
        except:
            return None
    
    # Calculate age
    df['Age'] = df[birthday_col_name].apply(calculate_age_safe)
    
    # Filter clients above 25
    clients_above_25 = df[df['Age'] > 25].copy()
    
    if not clients_above_25.empty:
        # Write headers
        headers = list(clients_above_25.columns)
        for col, header in enumerate(headers, 1):
            cell = ws_new.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Write data
        for row_idx, (_, row_data) in enumerate(clients_above_25.iterrows(), 2):
            for col_idx, value in enumerate(row_data, 1):
                # Handle datetime objects
                if isinstance(value, pd.Timestamp):
                    value = value.strftime('%Y-%m-%d')
                elif pd.isna(value):
                    value = ""
                ws_new.cell(row=row_idx, column=col_idx, value=value)
        
        # Apply table formatting
        table_ref = f"A1:{get_column_letter(len(headers))}{len(clients_above_25) + 1}"
        table = Table(displayName="ClientsAbove25Static", ref=table_ref)
        
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True
        )
        table.tableStyleInfo = style
        ws_new.add_table(table)
        
        # Auto-adjust column widths
        for column in ws_new.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_new.column_dimensions[column_letter].width = adjusted_width
    
    else:
        # No data found
        ws_new.cell(row=1, column=1, value="No clients above 25 years old found")
    
    return ws_new

def create_task2_worksheet(wb, df, renewal_col_name):
    """Create Task 2: Static worksheet for July and August 2025 renewals"""
    
    # Remove existing sheet if it exists
    if "Task 2 - July Aug Renewals" in wb.sheetnames:
        wb.remove(wb["Task 2 - July Aug Renewals"])
    
    # Create new worksheet
    ws_new = wb.create_sheet("Task 2 - July Aug Renewals")
    
    # Convert renewal dates to datetime
    df[renewal_col_name] = pd.to_datetime(df[renewal_col_name], errors='coerce')
    
    # Filter for July and August 2025
    july_2025 = df[(df[renewal_col_name].dt.year == 2025) & (df[renewal_col_name].dt.month == 7)]
    august_2025 = df[(df[renewal_col_name].dt.year == 2025) & (df[renewal_col_name].dt.month == 8)]
    
    # Write July 2025 section
    current_row = 1
    
    # July section header
    ws_new.cell(row=current_row, column=1, value="JULY 2025 RENEWALS")
    header_cell = ws_new.cell(row=current_row, column=1)
    header_cell.font = Font(bold=True, size=14)
    header_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    header_cell.font = Font(bold=True, color="FFFFFF")
    
    # Merge cells for header
    ws_new.merge_cells(f'A{current_row}:{get_column_letter(len(df.columns))}{current_row}')
    current_row += 1
    
    # Add July data
    if not july_2025.empty:
        # Headers
        for col, header in enumerate(df.columns, 1):
            cell = ws_new.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
        current_row += 1
        
        # Data rows
        for _, row_data in july_2025.iterrows():
            for col, value in enumerate(row_data, 1):
                ws_new.cell(row=current_row, column=col, value=value)
            current_row += 1
    else:
        ws_new.cell(row=current_row, column=1, value="No renewals found for July 2025")
        current_row += 1
    
    # Add spacing
    current_row += 2
    
    # August section header
    ws_new.cell(row=current_row, column=1, value="AUGUST 2025 RENEWALS")
    header_cell = ws_new.cell(row=current_row, column=1)
    header_cell.font = Font(bold=True, size=14)
    header_cell.fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
    header_cell.font = Font(bold=True, color="FFFFFF")
    
    # Merge cells for header
    ws_new.merge_cells(f'A{current_row}:{get_column_letter(len(df.columns))}{current_row}')
    current_row += 1
    
    # Add August data
    if not august_2025.empty:
        # Headers
        for col, header in enumerate(df.columns, 1):
            cell = ws_new.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E5F9F6", end_color="E5F9F6", fill_type="solid")
        current_row += 1
        
        # Data rows
        for _, row_data in august_2025.iterrows():
            for col, value in enumerate(row_data, 1):
                ws_new.cell(row=current_row, column=col, value=value)
            current_row += 1
    else:
        ws_new.cell(row=current_row, column=1, value="No renewals found for August 2025")
        current_row += 1
    
    # Auto-adjust column widths
    for column in ws_new.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_new.column_dimensions[column_letter].width = adjusted_width
    
    return ws_new, july_2025, august_2025

def create_enhanced_excel_with_both_tasks(uploaded_file, raw_data_sheet, birthday_col, renewal_col):
    """Create enhanced Excel file with both Task 1 and Task 2"""
    
    # Read the original file
    df = pd.read_excel(uploaded_file, sheet_name=raw_data_sheet)
    
    # Create temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        df.to_excel(tmp_file.name, sheet_name=raw_data_sheet, index=False)
        
        # Load workbook
        wb = load_workbook(tmp_file.name)
        
        # Create Task 1 worksheet
        task1_ws = create_task1_worksheet(wb, raw_data_sheet, birthday_col)
        
        # Create Task 2 worksheet
        task2_ws, july_data, august_data = create_task2_worksheet(wb, df, renewal_col)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Clean up temp file
        os.unlink(tmp_file.name)
        
        return output.getvalue(), july_data, august_data

if uploaded_file:
    try:
        # Read all sheets
        excel_file = pd.ExcelFile(uploaded_file)
        sheet_names = excel_file.sheet_names
        
        st.subheader("📋 Available Sheets:")
        st.write(f"Found sheets: {', '.join(sheet_names)}")
        
        # Let user select sheet as raw data
        raw_data_sheet = st.selectbox(
            "Select Raw Data sheet:",
            options=sheet_names,
            index=0
        )
        
        # Read selected sheet
        df = pd.read_excel(uploaded_file, sheet_name=raw_data_sheet)
        
        # Clean column names
        df.columns = df.columns.str.strip()
        
        # Show preview
        st.subheader("📊 Raw Data Preview:")
        st.dataframe(df.head())
        
        # Show all columns for manual selection if needed
        st.subheader("📋 All Available Columns:")
        st.write("Available columns in your data:")
        for i, col in enumerate(df.columns):
            st.write(f"{i+1}. **{col}** - Sample: {df[col].dropna().iloc[0] if not df[col].dropna().empty else 'No data'}")
        
        # Find potential date columns
        date_columns = find_date_columns(df)
        email_columns = find_email_columns(df)
        
        st.subheader("🔍 Detected Columns:")
        st.write(f"**Date columns found:** {', '.join(date_columns) if date_columns else 'None'}")
        st.write(f"**Email columns found:** {', '.join(email_columns) if email_columns else 'None'}")
        
        # Manual column selection with all columns as options
        st.subheader("🎯 Column Selection:")
        
        # For birthday - show all columns
        birthday_col = st.selectbox(
            "Select Birthday column (for Task 1 - age calculation):",
            options=df.columns.tolist(),
            index=df.columns.tolist().index('Birthday') if 'Birthday' in df.columns else 0
        )
        
        # For renewal date - prefer detected date columns
        renewal_options = date_columns if date_columns else df.columns.tolist()
        renewal_col = st.selectbox(
            "Select Renewal/Premium Date column (for Task 2):",
            options=renewal_options,
            index=renewal_options.index('Next Premium Date') if 'Next Premium Date' in renewal_options else 0
        )
        
        # Show sample data for selected columns
        st.subheader("🔍 Selected Columns Preview:")
        col1, col2 = st.columns(2)
        
        with col1:
            st.write(f"**Birthday Column ({birthday_col}):**")
            birthday_samples = df[birthday_col].dropna().head(5).tolist()
            for i, val in enumerate(birthday_samples, 1):
                st.write(f"  {i}. {val}")
        
        with col2:
            st.write(f"**Renewal Date Column ({renewal_col}):**")
            renewal_samples = df[renewal_col].dropna().head(5).tolist()
            for i, val in enumerate(renewal_samples, 1):
                st.write(f"  {i}. {val}")
        
        # Validate the selected columns
        try:
            # Test birthday column
            test_birthday = pd.to_datetime(df[birthday_col].dropna().head(5), errors='coerce', dayfirst=True)
            if test_birthday.isna().all():
                st.error(f"❌ Birthday column '{birthday_col}' does not contain valid dates. Please select a different column.")
                st.stop()
            else:
                st.success(f"✅ Birthday column '{birthday_col}' contains valid dates")
            
            # Test renewal column
            test_renewal = pd.to_datetime(df[renewal_col].dropna().head(5), errors='coerce', dayfirst=True)
            if test_renewal.isna().all():
                st.error(f"❌ Renewal column '{renewal_col}' does not contain valid dates. Please select a different column.")
                st.stop()
            else:
                st.success(f"✅ Renewal column '{renewal_col}' contains valid dates")
                
        except Exception as e:
            st.error(f"Error validating columns: {str(e)}")
            st.stop()
        if st.button("🚀 Generate Enhanced Excel with Both Tasks"):
            with st.spinner("Creating enhanced Excel file with both tasks..."):
                
                try:
                    enhanced_excel, july_data, august_data = create_enhanced_excel_with_both_tasks(
                        uploaded_file, raw_data_sheet, birthday_col, renewal_col
                    )
                    
                    if enhanced_excel:
                        st.success("✅ Enhanced Excel file created successfully!")
                        
                        # Show previews
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.subheader("📋 Task 1 Preview: Clients Above 25")
                            # Calculate age for preview
                            df[birthday_col] = pd.to_datetime(df[birthday_col], errors='coerce')
                            df['Age'] = df[birthday_col].apply(calculate_age)
                            clients_above_25 = df[df['Age'] > 25]
                            st.write(f"Total clients above 25: {len(clients_above_25)}")
                            if not clients_above_25.empty:
                                st.dataframe(clients_above_25[['Client Name', 'Age', 'Email'] if 'Client Name' in clients_above_25.columns and 'Email' in clients_above_25.columns else clients_above_25.head()])
                        
                        with col2:
                            st.subheader("📋 Task 2 Preview: July & August 2025")
                            st.write(f"July 2025 renewals: {len(july_data)}")
                            st.write(f"August 2025 renewals: {len(august_data)}")
                            
                            if not july_data.empty:
                                st.write("**July 2025 Sample:**")
                                st.dataframe(july_data[['Client Name', renewal_col, 'Email'] if 'Client Name' in july_data.columns and 'Email' in july_data.columns else july_data.head(3)])
                            
                            if not august_data.empty:
                                st.write("**August 2025 Sample:**")
                                st.dataframe(august_data[['Client Name', renewal_col, 'Email'] if 'Client Name' in august_data.columns and 'Email' in august_data.columns else august_data.head(3)])
                        
                        # Download button
                        st.download_button(
                            label="⬇️ Download Enhanced Excel File (Both Tasks)",
                            data=enhanced_excel,
                            file_name="enhanced_client_data_both_tasks.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
                        # Email extraction
                        st.subheader("📧 Email Lists for Policy Changes")
                        
                        if email_columns:
                            email_col = email_columns[0]
                            
                            # Task 1 emails
                            if not clients_above_25.empty:
                                task1_emails = clients_above_25[email_col].dropna().tolist()
                                st.write(f"**Task 1 - Clients Above 25 Emails ({len(task1_emails)} emails):**")
                                st.text_area("Copy these emails:", value="; ".join(task1_emails), height=100)
                            
                            # Task 2 emails
                            july_emails = july_data[email_col].dropna().tolist() if not july_data.empty else []
                            august_emails = august_data[email_col].dropna().tolist() if not august_data.empty else []
                            
                            if july_emails:
                                st.write(f"**July 2025 Renewal Emails ({len(july_emails)} emails):**")
                                st.text_area("Copy July emails:", value="; ".join(july_emails), height=100)
                            
                            if august_emails:
                                st.write(f"**August 2025 Renewal Emails ({len(august_emails)} emails):**")
                                st.text_area("Copy August emails:", value="; ".join(august_emails), height=100)
                        
                        st.subheader("🎉 Features Summary:")
                        st.write("""
                        **Task 1 - Auto-Updating Clients Above 25:**
                        ✅ Automatically updates when raw data changes
                        ✅ Dynamic age calculation using TODAY() function
                        ✅ Formula-based filtering for clients above 25
                        ✅ Professional table formatting
                        
                        **Task 2 - July & August 2025 Renewals:**
                        ✅ Static list (no auto-update needed)
                        ✅ Separate sections for July and August
                        ✅ Professional formatting with colors
                        ✅ Easy email extraction
                        
                        **General Features:**
                        ✅ Email lists easily extractable
                        ✅ Professional Excel formatting
                        ✅ Auto-adjusted column widths
                        ✅ Ready for policy change notifications
                        """)
                        
                        st.subheader("⚠️ Limitations:")
                        st.write("""
                        - Task 1 requires Excel 2016+ for full formula support
                        - Task 2 is static and won't update automatically
                        - Date format must be consistent in source data
                        - Email column detection is based on column name containing 'email'
                        """)
                        
                    else:
                        st.error("Failed to create enhanced Excel file")
                        
                except Exception as e:
                    st.error(f"Error creating enhanced Excel: {str(e)}")
                    st.write("Please check your data format and column selections.")
        
    except Exception as e:
        st.error(f"Error processing file: {str(e)}")
        st.write("Please check your file format and try again.")
        
else:
    st.info("Please upload an Excel file to begin analysis.")
    
    # Show expected format
    st.subheader("📝 Expected Data Format:")
    st.write("""
    Your Excel file should contain columns like:
    - **Client Name** (or Name)
    - **NRIC** (or ID)
    - **Birthday** (or DOB, Birth Date)
    - **Phone**
    - **Email**
    - **Policy Number**
    - **Policy Name**
    - **Next Premium Date** (or Renewal Date)
    """)
    
    # Sample data preview
    sample_data = pd.DataFrame({
        'S/No.': [1, 2, 3, 4],
        'Client Name': ['John Doe', 'Jane Smith', 'Bob Johnson', 'Alice Brown'],
        'NRIC': ['S1234567A', 'S7654321B', 'S1122334C', 'S9988776D'],
        'Birthday': ['1990-08-15', '1985-09-22', '1992-12-10', '2005-03-18'],
        'Phone': ['91234567', '98765432', '87654321', '81234567'],
        'Email': ['john@email.com', 'jane@email.com', 'bob@email.com', 'alice@email.com'],
        'Policy Number': ['POL001', 'POL002', 'POL003', 'POL004'],
        'Next Premium Date': ['2025-07-15', '2025-08-22', '2025-07-10', '2025-09-01']
    })
    st.dataframe(sample_data)
